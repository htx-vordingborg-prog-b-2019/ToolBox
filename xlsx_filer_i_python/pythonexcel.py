#Her importerer jeg to funktioner fra openpyxl biblioteket. Dette skal downloades først, hvilket kan gøres ved at skrive "pip install openpyxl" i din kommandoprompt.
from openpyxl import Workbook, load_workbook

#Her vil jeg oprette en xlsx-fil og skrive ting i den

wb1 = Workbook()
#Her bruges Workbook-funktionen til at oprette en tom excel-fil som gemmes i variablen wb1 (workbook1)
sheet1 = wb1.active
#Her sætter jeg en variabel til at indeholde det nuværende ark i siden, og siden at der kun er et ark i en ny fil er det den første
sheet1.title = 'ArkEksempel'
#Her navngiver jeg arket. Dette er ikke nødvendigt da arket ellers bare hedder Ark1 eller Sheet1, det er bare for at vise at man kan
sheet1['A1'] = "Monty Python Flying Circus rocks!"
#Her finder jeg feltet A1 i arket og sætter det til en tekst
sheet1['B3'] = "Marcus McCool Christiansen er den sejeste i hele klassen."
#Her finder jeg feltet B3 i arket og sætter det til en tekst
sheet1['A6'] = "2R>2Q."
#Her finder jeg feltet A6 i arket og sætter det til en tekst
wb1.save('eksempelKopi.xlsx')
#Her gemmer jeg filen under navnet eksempelKopi. Husk at gøre dette, da filen ellers ikke kan bruges, og husk .xlsx, da filen ellers ikke er en excel-fil.


#Her vil jeg åbne en xlsx-fil og læse den

wb2 = load_workbook('eksempelKopi.xlsx')
#Jeg bruger her funktionen load_workbook til at finde en fil i samme mappe som programmet med navnet eksempelKopi.xlsx. Det er filen vi lige har gemt, men jeg gemmer filen i en anden variabel for at vise at filen bliver loadet på ny.
sheet2 = wb2.active
#Her sætter jeg en ny variabel til at indeholde det nuværende ark i wb2.
print(sheet2['A1'].value+" "+sheet2['B3'].value+" "+sheet2['A6'].value)
#Her printer jeg værdierne i de tre felter A1, B3 og A6. Man kan også bruge disse værdier til andet end at printe dem.